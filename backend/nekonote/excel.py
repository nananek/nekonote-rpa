"""Excel and CSV operations for nekonote scripts.

Usage::

    from nekonote import excel

    data = excel.read("input.xlsx")
    excel.write("output.xlsx", data)
"""

from __future__ import annotations

import csv as _csv
import io
from pathlib import Path
from typing import Any

from nekonote.errors import FileNotFoundError as NkFileNotFoundError


def _require_file(path: str, action: str) -> Path:
    p = Path(path)
    if not p.is_file():
        nearby = [f.name for f in p.parent.iterdir()][:20] if p.parent.is_dir() else []
        raise NkFileNotFoundError(
            f"File not found: {path}",
            action=action,
            context={"path": str(p.resolve()), "nearby_files": nearby},
            suggestion=f"Files in {p.parent}: {', '.join(nearby[:5])}" if nearby else "",
        )
    return p


# ---------------------------------------------------------------------------
# Excel (.xlsx)
# ---------------------------------------------------------------------------


def read(
    path: str,
    *,
    sheet: str = "",
    header: bool = True,
) -> list[dict[str, Any]] | list[list[Any]]:
    """Read an Excel file.

    Returns list of dicts if *header* is True (first row = keys),
    otherwise list of lists.
    """
    from openpyxl import load_workbook

    _require_file(path, "excel.read")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    if header:
        headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
        return [dict(zip(headers, row)) for row in rows[1:]]
    return [list(row) for row in rows]


def read_cell(path: str, *, sheet: str = "", cell: str = "A1") -> Any:
    """Read a single cell value."""
    from openpyxl import load_workbook

    _require_file(path, "excel.read_cell")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    value = ws[cell].value
    wb.close()
    return value


def get_sheet_names(path: str) -> list[str]:
    """Return sheet names in the workbook."""
    from openpyxl import load_workbook

    _require_file(path, "excel.get_sheet_names")
    wb = load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def write(
    path: str,
    data: list[dict[str, Any]] | list[list[Any]],
    *,
    sheet: str = "Sheet1",
) -> str:
    """Write data to an Excel file (creates or overwrites).

    *data* can be list of dicts (keys become headers) or list of lists.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    if data and isinstance(data[0], dict):
        headers = list(data[0].keys())
        ws.append(headers)
        for row in data:
            ws.append([row.get(h) for h in headers])
    else:
        for row in data:
            ws.append(list(row))

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    return str(Path(path).resolve())


def write_cell(path: str, *, sheet: str = "", cell: str = "A1", value: Any) -> None:
    """Write a value to a single cell."""
    from openpyxl import load_workbook

    p = Path(path)
    if p.exists():
        wb = load_workbook(path)
    else:
        from openpyxl import Workbook

        wb = Workbook()

    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    ws[cell] = value
    wb.save(path)
    wb.close()


def append(
    path: str,
    rows: list[dict[str, Any]] | list[list[Any]],
    *,
    sheet: str = "",
) -> None:
    """Append rows to an existing Excel file."""
    from openpyxl import load_workbook

    _require_file(path, "excel.append")
    wb = load_workbook(path)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    for row in rows:
        if isinstance(row, dict):
            ws.append(list(row.values()))
        else:
            ws.append(list(row))

    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def read_csv(
    path: str,
    *,
    encoding: str = "utf-8",
    delimiter: str = ",",
    header: bool = True,
) -> list[dict[str, str]] | list[list[str]]:
    """Read a CSV file."""
    _require_file(path, "excel.read_csv")
    with open(path, "r", encoding=encoding, newline="") as f:
        if header:
            reader = _csv.DictReader(f, delimiter=delimiter)
            return [dict(row) for row in reader]
        reader = _csv.reader(f, delimiter=delimiter)
        return [list(row) for row in reader]


def write_csv(
    path: str,
    data: list[dict[str, Any]] | list[list[Any]],
    *,
    encoding: str = "utf-8",
    delimiter: str = ",",
) -> str:
    """Write data to a CSV file."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding=encoding, newline="") as f:
        if data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            writer = _csv.DictWriter(f, fieldnames=headers, delimiter=delimiter)
            writer.writeheader()
            writer.writerows(data)
        else:
            writer = _csv.writer(f, delimiter=delimiter)
            writer.writerows(data)
    return str(Path(path).resolve())
